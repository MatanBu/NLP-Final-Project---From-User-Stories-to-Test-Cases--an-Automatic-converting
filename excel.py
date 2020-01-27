from openpyxl import load_workbook

class Excel():
    
    USs = []        # { F# , name , description , Acceptance Criteria }
    USNames = []    # user stories names
    TCs = []        # {TC#, TC name , TC description }   
    TC_data = []
    
    def __init__(self, path):   
        """ 
        Load excel file and process its' sheets.
        Break down to:
            - USs: list of User Stories (tupled as mentioned above)
            - USNames: list of User Story names only
            - TCs: list of Test Cases
        """   
        wb = load_workbook(path)
        data_sheets_names = wb.sheetnames  
            
        # Get all the sheet names that hold User stories and Test Cases
        USs_sheets_names = []
        TCs_Sheets_names = []
              
        for sheet_name in data_sheets_names:
            if sheet_name.find("TCs") != -1:
                TCs_Sheets_names.append(sheet_name)
            if sheet_name.find(" F") != -1:
                USs_sheets_names.append(sheet_name)
            
        # get NLP data sheets     
        for name in USs_sheets_names:
            sheet = wb[name]
            for i in range(2,sheet.max_row+1):
                self.USs.append(((sheet.cell(i, 1).value),(sheet.cell(i, 2).value),(sheet.cell(i, 3).value),(sheet.cell(i, 4).value)))
            self.USNames.append((sheet.cell(2, 2).value))    
            
        for name in TCs_Sheets_names:
            sheet = wb[name]
            for i in range(2,sheet.max_row+1):
                self.TCs.append((name,(sheet.cell(i, 1).value),(sheet.cell(i, 2).value)))
        
        self.USNames = list(set(self.USNames))
    
    def get_US_names(self):
        return self.USNames
    
    def get_US_Des(self,US):
        """ Find corresponding Description to User Stories""" 
        #  fill AC combobox
        descriptions = []
        
        for a,b,c,d in self.USs:
            if b == US:
                descriptions.append(c)              
       
        return descriptions
    
    def get_Des_AC(self,Des):
        """ Find corresponding Acceptance Criteria to Description""" 
        #  fill AC combobox 
        ACs = []
        
        for a,b,c,d in self.USs:
            if c == Des:
                ACs.append(d)
                   
        return ACs
        
    def get_TCs(self):
        return self.TCs    
        
    def get_TC_data(self):
        return self.TC_data      
        
    def get_FNUM(self,US):
        """ Return feature number of user story"""
        for a,b,c,d in self.USs:
            if b == US:
                return a
            

