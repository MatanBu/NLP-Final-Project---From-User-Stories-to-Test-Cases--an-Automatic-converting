## Imports
import sys
from PyQt5.QtWidgets import QMainWindow, QApplication, QWidget, QPushButton, QLineEdit, QAction, QFileDialog, QVBoxLayout, QTableWidget, QTableWidgetItem, QPlainTextEdit
from PyQt5.QtGui import QIcon
from PyQt5.QtCore import pyqtSlot
from PyQt5.Qt import QLabel, QPoint, QComboBox, QFrame, QScrollArea
from PyQt5 import QtGui , QtCore

from nlp import tokenize, compare_sens
from operator import itemgetter
import openpyxl
from openpyxl import load_workbook

class App(QWidget):
    
    USs = []        # { F# , name , description , Acceptance Criteria }
    USNames = []    # user stories names
    TCs = []        # { TC name , TC description }    
    
    
    def __init__(self):
        super().__init__()
        self.title = 'NLP project.'
        self.left = 50
        self.top = 50
        self.width = 1500
        self.height = 800
        self.initUI()

        
    def initUI(self):
        self.setWindowTitle(self.title)
        self.setGeometry(self.left, self.top, self.width, self.height)
        app.setAttribute(QtCore.Qt.AA_EnableHighDpiScaling, True)
        
        # Create Label
        label = QLabel('Select User Story by name:', self)
        label.move(50,0)
        
        # Create Label
        label2 = QLabel('US Description + Acceptance Criteria:', self)
        label2.move(50,125)
        
        # Create Label
        label3 = QLabel('Choose Algorithm:', self)
        label3.move(50,600)
        
        # Create Label
        label4 = QLabel('Results:', self)
        label4.move(1000,50)
        
        # Create textbox user story
        self.textbox = QPlainTextEdit(self)
        self.textbox.move(50, 175)
        self.textbox.resize(500,300)
        
       
        """
        # Create textbox results
        self.textbox3 = QLineEdit(self)
        self.textbox3.move(1000, 75)
        self.textbox3.resize(400,600)
        self.textbox3.setDisabled(True)
        """
        
        self.tableWidget = QTableWidget()
        
        #self.tableWidget.resize(400,600)
        
        # Create textbox tokenize
        
        self.textbox4 = QLineEdit(self)
        self.textbox4.move(50, 700)
        self.textbox4.resize(280,40)
        #self.textbox4.setDisabled(True)
        """
        self.scrollArea = QScrollArea(self)
        self.scrollArea.setWidgetResizable(True)
        self.scrollArea.move(50,400)
        self.scrollArea.resize(280,40)
        self.scrollArea.setWidget(self.textbox4)
        """     
        button = QPushButton('Start', self)
        button.move(150,750) 
        button.clicked.connect(self.on_click_start)
        
        button2 = QPushButton('Tokenize', self)
        button2.move(50, 625)
        button2.clicked.connect(self.on_click_tokenize)
        
        button3 = QPushButton('Load excel', self)
        button3.move(50, 750)
        button3.clicked.connect(self.on_click_load)
        
        button4 = QPushButton('Choose US', self)
        button4.move(550, 29)
        button4.clicked.connect(self.on_click_choose_US)
        
        button5 = QPushButton('Choose AC', self)
        button5.move(550, 149)
        button5.clicked.connect(self.on_click_choose_AC)
        
        button6 = QPushButton('Clear', self)
        button6.move(250,750) 
        button6.clicked.connect(self.on_click_clear)
        
        # algorithm combobox
        self.comboBox = QComboBox(self)
        algorithms = ['cosine','other']
        self.comboBox.addItems(algorithms)
        self.comboBox.move(150, 625)
        
        # US combobox
        self.comboBox2 = QComboBox(self)
        self.comboBox2.move(50, 25)
        self.comboBox2.resize(500,20)
        #action_relect = lambda: self.on_click_reselect_us(self)
        self.comboBox2.currentIndexChanged.connect(self.on_click_reselect_us)
        
        # AC combobox
        self.comboBox3 = QComboBox(self)
        self.comboBox3.move(50, 150)
        self.comboBox3.resize(500,20)
        
        
        self.show()
        
    @pyqtSlot()
    def on_click_reselect_us(self):
        
        self.textbox.clear() 
        self.comboBox3.clear()
        
            
    @pyqtSlot()
    def on_click_clear(self):
        
        self.textbox.clear()
        self.textbox3.clear()
        self.textbox4.clear()
        self.comboBox.clear()
        self.comboBox2.clear()
        self.comboBox3.clear()
        
    @pyqtSlot()
    def on_click_start(self):

        us = self.textbox2.text()
        results = []
        for tc in self.TCs:
            tmp = tc[0] + tc[1]
            results.append({'match_sentence':tmp,'score':compare_sens(us, tmp)})
        
        i = 0
        self.tableWidget.setSortingEnabled(True)
        self.tableWidget.setRowCount(100)
        self.tableWidget.setColumnCount(2)
        results = sorted(results, key=itemgetter('score'), reverse=True)
        
        for res in results:
            #self.tableWidget.insertRow(str(res)+str1)
            #self.tableWidget.setItem(i,0, QTableWidgetItem(str(res[1])))
            self.tableWidget.setItem(i,0, QTableWidgetItem(str(res)))
            i+=1
            #self.textbox3.insert(str(res)+str1
        self.tableWidget.move(1000,75)
        self.tableWidget.show()
        
    
    @pyqtSlot()
    def on_click_choose_US(self):
        
        # set User Story over textbox so we can edit if needed
        text = self.comboBox2.currentText()
        self.textbox.setPlainText(text)
        self.textbox.show()
        
        #  fill AC combobox
        descriptions = []
        ACs = []
        tmp = []
        us = self.comboBox2.currentText()
        
        for a,b,c,d in self.USs:
            if b == us:
                descriptions.append(c)
                ACs.append(d)
                tmp.append(c+d)
        self.comboBox3.addItems(tmp)
    
    @pyqtSlot()
    def on_click_choose_AC(self):

        # set AC over textbox so we can edit if needed
        text = self.comboBox3.currentText()
        self.textbox.paste()
        self.textbox.show()
                
    @pyqtSlot()
    def on_click_tokenize(self):

        text = self.textbox.text()
        tok = str( tokenize(text) )
        self.textbox4.setText(tok)
     
        
    @pyqtSlot()
    def on_click_load(self):
        
        # Assign spreadsheet filename to `file`
        file = self.openFileNameDialog()
    
        # Load in the workbook
        wb = load_workbook(file)

        # Get sheet names
        data_sheets_names = wb.sheetnames
        print(wb.sheetnames)
    
        # Get all the sheet names that hold User stories and Test Cases
        USs_sheets_names = []
        TCs_Sheets_names = []
        
        for sheet_name in data_sheets_names:
            if sheet_name.find("TCs") != -1:
                TCs_Sheets_names.append(sheet_name)
            if sheet_name.find(" F") != -1:
                USs_sheets_names.append(sheet_name)
        
        # get NLP data sheets     
        for name in USs_sheets_names:
            sheet = wb[name]
            for i in range(2,sheet.max_row+1):
                self.USs.append(((sheet.cell(i, 1).value),(sheet.cell(i, 2).value),(sheet.cell(i, 3).value),(sheet.cell(i, 4).value)))
            self.USNames.append((sheet.cell(2, 2).value))    
        
        for name in TCs_Sheets_names:
            sheet = wb[name]
            for i in range(2,sheet.max_row+1):
                self.TCs.append((name,(sheet.cell(i, 1).value),(sheet.cell(i, 2).value)))
                
        #remove doubles
        self.USNames = list(set(self.USNames))     
        self.comboBox2.addItems(self.USNames)
        
        print(self.USNames)
    
    def openFileNameDialog(self):    
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        fileName, _ = QFileDialog.getOpenFileName(self,"QFileDialog.getOpenFileName()", "","All Files (*);;Excel Files (*.xlsx)", options=options)
        if fileName:
            print(fileName)    
        return fileName    
            
if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = App()
    sys.exit(app.exec_())
